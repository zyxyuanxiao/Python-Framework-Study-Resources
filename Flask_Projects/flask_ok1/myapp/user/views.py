#!/usr/bin/env python
# encoding: utf-8

from myapp.user import user
from exts import db
from myapp.models import Match, Gambler, Monitor, Article, IpList, Biao1, Biao2
from flask_login import login_user, logout_user, current_user
from flask import session, g, redirect, url_for, request, render_template

from ctrl.register import *
from ctrl.model_class_serialize import *

import datetime
import time
from functools import wraps

now_time = datetime.datetime.now()


# Ex对象
class ListObj:
    def __init__(self, jy, num):
        self.jy = jy
        self.num = num

    def thisobj(self):
        obj_idr = {self.jy: self.num}
        return obj_idr


# 连续写入
class RecordLogs(object):
    def __init__(self, msg, p):
        self.msg = msg
        self.p = p
        with open(p, 'a+') as f:
            f.write(msg)


p = '/root/myproject/phone_msg.txt'
p1 = '/Users/yangyuexiong/Desktop/flask_ok1/phone_msg.txt'

'''_____________________________________________________________'''


def user_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('user.lien'))
        return f(*args, **kwargs)

    return decorated_function


# 拦截调试
@user.route('/test')
@user_required
def test():
    return '已经登录'


@user.route('/lien')
def lien():
    return '在这这个页面登录'


# # flask-login
# @login_manager.user_loader
# def load_user(gambler_id):
#     return Gambler.query.get(gambler_id)


@user.before_request
def before_request():
    g.gambler = current_user


@user.context_processor
def my_context_processor():
    gambler_id = session.get('gambler_id')
    if gambler_id:
        gambler = Gambler.query.filter(Gambler.id == gambler_id).first()
        return {'gambler': gambler}
    return {}


# app注册
@user.route('/register', methods=['POST'])
def app_register():
    username = request.form.get('username')
    password = request.form.get('password')
    password2 = request.form.get('password2')
    print(username, password, password2)

    gambler = Gambler.query.filter_by(username=username).first()
    print(gambler)
    if gambler:
        return ApiResult().formattingData(status=0, msg='用户名已经被注册!', data=[])

    if username == '':
        return ApiResult().formattingData(status=1, msg='手机号不能为空!', data=[])

    if len(username) <= 6:
        return ApiResult().formattingData(status=2, msg='用户名长度应大于6位字符或数字!', data=[])

    if password == '':
        return ApiResult().formattingData(status=3, msg='密码不能为空!', data=[])

    if password != password2:
        return ApiResult().formattingData(status=4, msg='两次输入的密码不一致!', data=[])

    if len(password) < 6 or len(password) > 14:
        return ApiResult().formattingData(status=5, msg='密码少于6位或多于14!', data=[])

    gambler = Gambler(username, password)
    db.session.add(gambler)
    db.session.commit()

    return ApiResult().formattingData(status=200, msg='注册成功!', data={
        "user_id": gambler.id,
        "mobile": gambler.username,
        "mobile_encode": '',
        "user_levle": gambler.level,
        "last_login_ip": '',
        "login_time": '',
        "nickname": '',
        'username': gambler.username,

        "avatar_original": {
            "life_time": '',
            "login_encode": 'login_encode',
        },

        "api_token": {"life_time": 'life_time',
                      },
    })


# app登陆
@user.route('/login', methods=['GET', 'POST'])
def app_login():
    # username = request.args.get('username')
    # password = request.args.get('password')
    username = request.form.get('username')
    password = request.form.get('password')
    # print(username, password)
    gambler = Gambler.query.filter_by(username=username).first()
    # print(gambler)
    p = gambler.check_passwodr(password)
    # p = True

    if not gambler:
        return ApiResult().formattingData(status=404, msg='手机号未注册', data={})

    if username == '':
        return ApiResult().formattingData(status=403, msg='用户名-不能为空', data={})

    if password == '':
        return ApiResult().formattingData(status=402, msg='密码-不能为空', data={})

    if not p:
        return ApiResult().formattingData(status=401, msg='密码-错误', data={})

    # session['gambler_id'] = gambler.id
    session.permanent = True
    login_user(user=gambler)

    u = Monitor(mobile=username)
    print('>>', u)

    # m = hashlib.md5()
    # m.update(username.encode(encoding="utf-8"))
    # m.update(password.encode(encoding="utf-8"))
    # m.update(str(int(time.time())).encode(encoding="utf-8"))
    # token = m.hexdigest()
    # print(token)
    #
    # r.hmset('user:%s' % user.mobile, {'token': token,
    #                                   'nickname': user.username,
    #                                   'app_online': 1})
    # r.set('token:%s' % token, user.mobile)
    # r.expire('token:%s' % token, 3600 * 24 * 30)

    # tk = session.pop('api_token', None)
    # if tk == None:
    #     return ApiResult().formattingData(status=-1, msg='none!', data=[])
    #
    # api_token = request.form.get('api_token')
    # if api_token != tk:
    #     return ApiResult().formattingData(status=0, msg='ok!', data=[])
    #
    # tk = str(uuid.uuid1())
    # r.set('login-' + mobile, tk, ex=3600 * 24 * 30)
    # r.set('login-' + mobile, tk, ex=30)

    m = Monitor.query.filter_by(mobile=username).first()
    if not m:
        m1 = Monitor(username)
        db.session.add(m1)
        db.session.commit()
    else:
        m2 = Monitor(mobile=request.form.get('username'))
        db.session.add(m2)
        db.session.commit()

    return ApiResult().formattingData(status=200, msg='登录成功', data={
        'user_id': gambler.id,
        'mobile': gambler.username,
        'user_levle': gambler.level,
        'last_login_time': '',
        'login_time': now_time,
        'nickname': gambler.nickname,
        'token': '',
        'login_encode': {
            'life_time': '',
            'login_encode': '',
        }
    })


# app退出
@user.route('/logout')
@user_required
def app_logout():
    session.clear()
    logout_user()
    # return redirect(url_for('app_index'))
    return redirect('user/app_index')


# 1.app首页-比赛列表
@user.route('/app_index', methods=['GET'])
def app_index():
    m = Match.query.order_by(Match.vs_code).all()
    arr = []
    for i in m:
        a = row2dict(i)
        arr.append(a)
    return ApiResult().formattingData(status=1, msg='', data=arr)


# app原创文章
@user.route('/app_article', methods=['GET'])
def article():
    m = Article.query.order_by(-Article.cea_time).filter(Article.state == 1).all()
    arr = []
    for i in m:
        a = row2dict(i)
        arr.append(a)
    return ApiResult().formattingData(status=1, msg='', data=arr)


# app转载文章
@user.route('/app_article_2', methods=['GET'])
def r_article():
    m = Article.query.order_by(-Article.cea_time).filter(Article.state == 2).all()
    arr = []
    for i in m:
        a = row2dict(i)
        arr.append(a)
    return ApiResult().formattingData(status=1, msg='', data=arr)


'___________________Other operations____________________'


# 定时更新ip
@user.route('/get_ip', methods=['GET'])
def get_ip():
    from qqwry import QQwry
    import datetime
    from myapp.models import IpList
    from qqwry import updateQQwry

    nowTime = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    print(nowTime)
    print('-------------开始时间------------')
    q = QQwry()
    q.load_file('/Users/yangyuexiong/Desktop/www.NewXing.com/qqwry.dat')
    arr = []
    ips = IpList.query.all()
    for i in ips:
        result = q.lookup(i.ip)
        print(result)
        x = result[0]
        y = result[1]
        # updata_ip = IpList.query.filter(ip=i.ip).update(
        #     province=x[0:3],
        #     city=x[3:6],
        #     area=x[6:9],
        #     operator=y, )

        updata_ip = IpList.query.filter_by(ip=i.ip).first()
        updata_ip.province = x[0:3]
        updata_ip.city = x[3:6]
        updata_ip.area = x[6:9]
        updata_ip.operator = y
        db.session.commit()

        a = row2dict(i)
        arr.append(a)
    print('--------------结束时间-----------')
    print(nowTime)
    return ApiResult().formattingData(status=1, msg='', data=arr)


# 提取excel存储mysql
@user.route('/get_ex', methods=['GET'])
def get_ex():
    import openpyxl
    file_2007 = '/Users/yangyuexiong/Desktop/时光姬销售数据.xlsx'

    wb = openpyxl.load_workbook(file_2007)
    sheet = wb['回传失败的码']
    for i in range(1, 547):
        x = sheet.cell(row=i, column=7).value
        y = sheet.cell(row=i, column=8).value
        l_obj = ListObj(x, y).thisobj()
        s = Biao1(jy=ListObj(x, y).jy, num=ListObj(x, y).num)
        db.session.add(s)
        db.session.commit()
        print(l_obj)
    print('写入完毕\n___________________________________')

    wb2 = openpyxl.load_workbook(file_2007)
    sheet2 = wb2['蜂巢发出去的订单（原始单号）']
    for i in range(1, 404):
        x = sheet2.cell(row=i, column=1).value
        y = sheet2.cell(row=i, column=2).value
        l_obj2 = ListObj(x, y).thisobj()
        s2 = Biao2(jy=ListObj(x, y).jy, num=ListObj(x, y).num)
        db.session.add(s2)
        db.session.commit()
        print(l_obj2)
    print('写入完毕\n___________________________________')
    return ApiResult().formattingData(status=1, msg='', data={})


@user.route('/get_iphone_msg')
def get_phone():
    return render_template('getphone.html')


@user.route('/okccc', methods=['POST'])
def oklala():
    v = request.form.get('version')
    u = request.form.get('u')
    os = request.form.get('os')
    m = ('名称:%s \n手机型号:%s \n版本号:%s\n\n') % (u, os, v)
    print(m)
    RecordLogs(m, p1)
    # print(v)
    # print(u)
    # print(os)
    return ApiResult().formattingData(status=200, msg='', data={
        'version': v,
        'nikename': u,
        'equipment': os
    })
